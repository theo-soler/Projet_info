# -*- coding: utf-8 -*-
def main():
    print('Starting import')
    from bs4 import BeautifulSoup
    import requests
    import pandas as pd
    import io
    import codecs
    import numpy as np
    import matplotlib.pyplot as plt
    from pathlib import Path
    import openpyxl
    print('Starting import : done')


    # Years we want
    li_annees_voulues = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]


    # website to webscrap
    urla = 'www.boursorama.com'


    # on recupere les liens des differentes pages de boursorama dont on aura besoin
    li_url_taux_de_change = []
    li_url_alu_alloy = []
    li_url_copper = []

    for an in li_annees_voulues:
        li_url_taux_de_change.append('www.boursorama.com/_formulaire-periode/?symbol=3fUSD_EUR&historic_search%5BstartDate%5D=01/01/'+ str(an) + '&historic_search%5Bduration%5D=1Y&historic_search%5Bperiod%5D=1')
        li_url_alu_alloy.append('www.boursorama.com/_formulaire-periode/?symbol=7xAAUSD15M&historic_search%5BstartDate%5D=01/01/'+ str(an) + '&historic_search%5Bduration%5D=1Y&historic_search%5Bperiod%5D=1')
        li_url_copper.append('www.boursorama.com/_formulaire-periode/?symbol=7xCAUSD&historic_search%5BstartDate%5D=01/01/'+ str(an) + '&historic_search%5Bduration%5D=1Y&historic_search%5Bperiod%5D=1')

    #on cree un dictionnaire pour l'aluminium dont les clés sont les dates et les valeurs du cours à cette date
    data_alu_alloy = {}

    print('Begin collecting data for aluminium')
    for url, year in zip(li_url_alu_alloy, li_annees_voulues) :
        r = requests.get('https://' + url)
        soup = BeautifulSoup(r.content, 'html.parser')
        diff_pages = soup.select('a')

        liens = []
        for el in diff_pages:
            liens.append(el['href'])
        print('Begin collecting data for aluminium in ' + str(year))
        for link in liens:
            rbis = requests.get('https://' + urla + link)
            soup = BeautifulSoup(rbis.content, 'html.parser')
            rows2 = soup.select('tbody tr td')
            for i in range(int(len(rows2)/6)-1):
                data_alu_alloy[rows2[6*i].text.strip()] = rows2[6*i + 1].text.strip()
        print('Collecting data for aluminium in ' + str(year) +': done')
    print('Collecting data for aluminium : done')
    # data_alu_alloy


    #on cree un dictionnaire pour le taux de change dont les clés sont les dates et les valeurs le taux de change à la date correspondante
    data_taux_de_change = {}

    print('Begin collecting data for exchange rate')
    for url, year in zip(li_url_taux_de_change, li_annees_voulues):
        r = requests.get('https://' + url)
        soup = BeautifulSoup(r.content, 'html.parser')
        diff_pages = soup.select('a')

        liens = []
        for el in diff_pages:
            liens.append(el['href'])
        print('Begin collecting data for exchange rate in ' + str(year))
        for link in liens:
            rbis = requests.get('https://' + urla + link)
            soup = BeautifulSoup(rbis.content, 'html.parser')
            rows2 = soup.select('tbody tr td')
            for i in range (int(len(rows2)/6)-1):
                data_taux_de_change[rows2[6*i].text.strip()] = rows2[6*i + 1].text.strip()
        print('Collecting data for exchange rate in ' + str(year) +': done')
    print('Collecting data for exchange rate : done')
    # data_taux_de_change


    #on cree un dictionnaire pour le taux de change dont les clés sont les dates et les valeurs le taux de change à la date correspondante
    data_copper = {}

    print('Begin collecting data for copper')
    for url, year in zip(li_url_copper, li_annees_voulues):
        r = requests.get('https://' + url)
        soup = BeautifulSoup(r.content, 'html.parser')
        diff_pages = soup.select('a')

        liens = []
        for el in diff_pages:
            liens.append(el['href'])
        print('Begin collecting data for copper in ' + str(year))
        for link in liens:
            rbis = requests.get('https://' + urla + link)
            soup = BeautifulSoup(rbis.content, 'html.parser')
            rows2 = soup.select('tbody tr td')
            for i in range (int(len(rows2)/6)-1):
                data_copper[rows2[6*i].text.strip()] = rows2[6*i + 1].text.strip()
        print('Collecting data for copper in ' + str(year) +': done')
    print('Collecting data for copper : done')
    # data_copper


    #on récupère le Brent csv sur le site
    print('Begin collecting csv for brent')
    URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?bgcolor=%23e1e9f0&chart_type=line&drp=0&fo=open%20sans&graph_bgcolor=%23ffffff&height=450&mode=fred&recession_bars=off&txtcolor=%23444444&ts=12&tts=12&width=1168&nt=0&thu=0&trc=0&show_legend=yes&show_axis_titles=yes&show_tooltip=yes&id=DCOILBRENTEU&scale=left&cosd=2017-05-23&coed=2022-05-23&line_color=%234572a7&link_values=false&line_style=solid&mark_type=none&mw=3&lw=2&ost=-99999&oet=99999&mma=0&fml=a&fq=Daily&fam=avg&fgst=lin&fgsnd=2020-02-01&line_index=1&transformation=lin&vintage_date=2022-05-27&revision_date=2022-05-27&nd=1987-05-20"
    response = requests.get(URL)
    open("data_extraite/brent_europe.csv", "wb").write(response.content)
    brent_data = response.content.decode('utf8')

    #on met en forme le jeu de données (format des valeurs et des dates)

    a = brent_data.replace('-', '/')
    a.replace('\n', ',')

    data = io.StringIO(a)
    df = pd.read_csv(data, sep=",")
    df_brent = df.rename(columns = {'DATE' : 'Date'})

    df_brent['Date'] = pd.to_datetime(df_brent['Date'], errors = 'coerce')
    df_brent['Date'] = df_brent['Date'].dt.strftime('%d/%m/%Y')
    df_brent_vrai = df_brent.set_index('Date')

    df_brent_vrai['DCOILBRENTEU'] = df_brent_vrai['DCOILBRENTEU'].str.replace(' ', '')
    df_brent_vrai['DCOILBRENTEU'] = pd.to_numeric(df_brent_vrai['DCOILBRENTEU'], errors='coerce')
    df_brent_vrai.columns = ['Brent (dollar per barrel)']
    print('Collecting csv for brent : done')


    # Fuel Fr
    print('Begin collecting data for fuel_FR')
    data_fuel_fr = {}
    r = requests.get('https://' + 'www.cnr.fr/espaces/13/indicateurs/43?noContext=1')
    soup = BeautifulSoup(r.content, 'html.parser')
    rows_value = soup.select('tr td span')
    rows_date = soup.select('tr th')
    li_dates = []
    for el in rows_date[1:]:
        li_dates.append(el.text.strip())
    data_fuel_fr = {}

    #ici les données sont arrondies au mois près sur le site : on complète avec 
    #des séries de données constantes (égale à la valeur mensuelle indiquée) sur chaque mois

    #on normalise le format des dates (JJ/MM/AAAA)

    for an in li_dates:
        print('Begin collecting data for fuel_FR in ' + str(an))
        for mois in range(1, 13) :
            for k in range(1,32):
                if k<=9:
                    if mois <= 9:
                        data_fuel_fr['0' + str(k) + '/0' + str(mois) + '/' + an] = rows_value[(mois-1)*23 + int(an)-2000].text.strip().replace(',', '.')
                    else:
                        data_fuel_fr['0'+str(k) + '/' + str(mois) + '/' + an] = rows_value[(mois-1)*23 + int(an)-2000].text.strip().replace(',', '.')
                else :
                    if mois <=9:
                        data_fuel_fr[str(k) + '/0' + str(mois) + '/' + an] = rows_value[(mois-1)*23 + int(an)-2000].text.strip().replace(',', '.')
                    else:
                        data_fuel_fr[str(k) + '/' + str(mois) + '/' + an] = rows_value[(mois-1)*23 + int(an)-2000].text.strip().replace(',', '.')
        print('Collecting data for fuel_FR in ' + str(an) +': done')
    print('Collecting data for fuel_FR : done')


    # Fuel All
    print('Begin collecting data for fuel_ALL')
    data_fuel_all = {}
    li_dates = []
    li_valeurs = []

    for an in li_annees_voulues:
        print('Begin collecting data for fuel_ALL in ' + str(an))
        r = requests.get('https://gasoline-germany.com/statistik.phtml?o=7&display=diesel' + str(an))
        soup = BeautifulSoup(r.content, 'html.parser')
        rows_value = soup.select('tr td font b')
        rows_date = soup.select('tr td font')
        li_valeurs.extend(rows_value[1:len(rows_value) - 1])
        for mois in range(len(rows_value)-1):
            for jour in range(1, 31):
                if mois < 10:    
                    if jour < 10:
                        data_fuel_all['0' + str(jour) + '/' + '0' + str(mois+1) + '/' + str(an)] = rows_value[mois].text[:5]
                    else :
                        data_fuel_all[str(jour) + '/' + '0' + str(mois+1) + '/' + str(an)] = rows_value[mois].text[:5]
                else : 
                    if jour < 10:
                        data_fuel_all['0' + str(jour) + '/' + str(mois+1) + '/' + str(an)] = rows_value[mois].text[:5]
                    else :
                        data_fuel_all[str(jour) + '/' + str(mois+1) + '/' + str(an)] = rows_value[mois].text[:5]
        print('Collecting data for fuel_ALL in ' + str(an) +': done')
    print('Collecting data for fuel_ALL : done')


    # Coils
    print('Begin collecting data for coils')
    file = codecs.open('Data/data_acier/MEPS-Europe Hot Dipped Galvanised Coil Price Forecast.html')
    r = bytes(file.read(), 'utf-8')
    soup = BeautifulSoup(r, 'html.parser')
    rows_date = soup.select('tr td b')[3:]
    rows_value = soup.select('tr td.price')

    dic_coils = {}

    #permet de normaliser le format des dates (JJ/MM/AAAA)

    for k in range(len(rows_date)):
        for i in range(1,32):
            if i <= 9:
                if (5 + k)%12 + 1 < 10:
                    dic_coils['0' + str(i) + '/0' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
                else :
                    dic_coils['0' + str(i) + '/' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
            else :
                if (5 + k)%12 + 1 < 10:
                    dic_coils[str(i) + '/0' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
                else :
                    dic_coils[str(i) + '/' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
    print('Collecting data for coils : done')


    # Beams
    print('Begin collecting data for beams')
    file = codecs.open('Data/data_acier/MEPS-Europe Sections & Beams Price Forecast.html')
    r = bytes(file.read(), 'utf-8')
    soup = BeautifulSoup(r, 'html.parser')
    rows_date = soup.select('tr td b')[3:]
    rows_value = soup.select('tr td.price')

    dic_beams = {}

    #permet de normaliser le format des dates (JJ/MM/AAAA)

    for k in range(len(rows_date)):
        for i in range(1,32):
            if i <= 9:
                if (5 + k)%12 + 1 < 10:
                    dic_beams['0' + str(i) + '/0' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
                else :
                    dic_beams['0' + str(i) + '/' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
            else :
                if (5 + k)%12 + 1 < 10:
                    dic_beams[str(i) + '/0' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
                else :
                    dic_beams[str(i) + '/' + str((5 + k)%12 + 1) + '/' + rows_date[k].text[4:]] = float(rows_value[k].text)
    print('Collecting data for beams : done')


    #convertit le dictionnaire data en dataframe en convertissant les valeurs par des float
    print('Converting values into float')
    df_alu_alloy = pd.DataFrame.from_dict(data_alu_alloy, orient = 'index', columns = ['aluminium (USD)'])
    df_alu_alloy['aluminium (USD)'] = df_alu_alloy['aluminium (USD)'].str.replace(' ', '')
    df_alu_alloy['aluminium (USD)'] = pd.to_numeric(df_alu_alloy['aluminium (USD)'], errors='coerce')

    df_taux_de_change = pd.DataFrame.from_dict(data_taux_de_change, orient = 'index', columns = ['USD/EUROS'])
    df_taux_de_change['USD/EUROS'] = pd.to_numeric(df_taux_de_change['USD/EUROS'], errors='coerce')

    df_copper = pd.DataFrame.from_dict(data_copper, orient = 'index', columns = ['copper'])
    df_copper['copper'] = df_copper['copper'].str.replace(' ', '')
    df_copper['copper'] = pd.to_numeric(df_copper['copper'], errors='coerce')

    df_fuel_fr = pd.DataFrame.from_dict(data_fuel_fr, orient = 'index', columns = ['fuel fr EUR/L'])
    df_fuel_fr['fuel fr EUR/L'] = df_fuel_fr['fuel fr EUR/L'].str.replace(' ', '')
    df_fuel_fr['fuel fr EUR/L'] = pd.to_numeric(df_fuel_fr['fuel fr EUR/L'], errors='coerce')

    df_coils = pd.DataFrame.from_dict(dic_coils, orient = 'index', columns = ['coils EUR/ton'])

    df_beams = pd.DataFrame.from_dict(dic_beams, orient = 'index', columns = ['beams EUR/ton'])

    df_fuel_all = pd.DataFrame.from_dict(data_fuel_all, orient = 'index', columns = ['fuel all EUR/L'])
    df_fuel_all['fuel all EUR/L'] = df_fuel_all['fuel all EUR/L'].str.replace(',', '.')
    df_fuel_all['fuel all EUR/L'] = pd.to_numeric(df_fuel_all['fuel all EUR/L'], errors='coerce')
    print('Converting values into float : done')


    #cree un fichier csv avec en index les dates et en colonne les cours aux différents moments
    # print('Creating csv for each data')
    # df_alu_alloy.to_csv (r'data_extraite/Alu_alloy_LME.csv', index = True, header=True)
    # df_taux_de_change.to_csv (r'data_extraite/Taux_echange.csv', index = True, header=True)
    # df_copper.to_csv (r'data_extraite/copper_LME.csv', index = True, header=True)
    # df_fuel_fr.to_csv(r'data_extraite/fuel_fr.csv', index = True, header=True)
    # df_coils.to_csv(r'data_extraite/coils.csv', index = True, header=True)
    # df_beams.to_csv(r'data_extraite/beams.csv', index=True, header = True)
    # df_fuel_all.to_csv(r'data_extraite/fuel_all.csv', index=True, header = True)
    # print('Creating csv for each data : done')


    #on réunit toutes les données
    print('Gathering data into one csv')
    tab_f = df_alu_alloy.join(df_taux_de_change).join(df_copper)
    tab_f = tab_f.join(df_brent_vrai)
    tab_f = tab_f.join(df_fuel_fr)
    tab_f = tab_f.join(df_coils)
    tab_f = tab_f.join(df_beams)
    tab_f = tab_f.join(df_fuel_all)
    print('Gathering data into one csv : done')


    #conversion en euros
    print('Converting values into euros')
    tab_f['aluminium (USD)'] = tab_f['aluminium (USD)']*tab_f['USD/EUROS']
    tab_f['Brent (dollar per barrel)'] = tab_f['Brent (dollar per barrel)']*tab_f['USD/EUROS']
    tab_f['copper'] = tab_f['copper']* tab_f['USD/EUROS']
    tab_f.columns = ['aluminium (EUR)', 'USD/EUROS', 'copper', 'Brent (EUR/barrel)', 'fuel fr EUR/L', 'coils EUR/ton', 'beams EUR/ton', 'fuel all EUR/L']
    print('Converting values into euros : done')


    #on enregistre les données selon un format plus pratique pour la suite
    print('Saving data')
    tab_f['Date'] = tab_f.index
    tab_melted = pd.melt(tab_f, id_vars=['Date'])
    tab_melted.to_csv(r'data_extraite/tableau_brut.csv', index = True, header = True)
    print('Saving data : done')


    # Tab by month
    print('Gathering data by month')
    tab = tab_melted
    tab['Date'] = pd.to_datetime(tab['Date'], format="%d/%m/%Y")
    tab = tab.set_index('Date')
    tab = np.round(tab.groupby(['variable', pd.Grouper(freq="M")], as_index=True)['value'].mean(), decimals = 2)
    tab = tab.reset_index()
    print('Gathering data by month : done')


    # Fill NaN
    print('Filling NaN')
    tab = tab.groupby('variable').apply(lambda x: x.ffill().bfill())
    print('Filling NaN : done')


    # Saving tab by month
    print('Saving data in tableau_final.csv')
    tab['Date'] = pd.to_datetime(tab.Date)
    tab['Date'].dt.strftime('%m/%Y')
    tab.to_csv(r'data_extraite/tableau_final.csv', index = True, header = True)
    print('Saving data in tableau_final.csv : done')


    # Saving plot
    print('Saving plot in results')
    for elem in tab['variable'].unique():
        tab_to_plot = tab[tab['variable'] == elem].plot(x='Date', y='value', legend=False)
        plt.ylabel(elem)
        output_file = Path('data_extraite/' + elem.replace("/", "-") + '.png')
        output_file.parent.mkdir(exist_ok=True, parents=True)
        plt.savefig('data_extraite/' + elem.replace("/", "-") + '.png')
    print('Saving plot in results : done')
    
    #importing prices
    #fenetres
    print('Importing prices for material : windows')
    wb_fenetre = openpyxl.load_workbook('Data/Extracts prix/Price Development Material Fenetre.xlsx')
    sheet_fenetre = wb_fenetre.active

    max_col_fenetre = sheet_fenetre.max_column
    dico_fenetre = {}
    for i in range(2, max_col_fenetre+1):
        dico_fenetre[sheet_fenetre.cell(row = 7, column = i).value] = sheet_fenetre.cell(row = 8, column = i).value 


    liste = list(dico_fenetre.items())
    dico_a_ajouter = {} 
    for el in range(1, len(dico_fenetre.items())) :
        date, valeur = liste[el]
        date_avant, valeur_avant = liste[el-1]
        if int(date[-2:]) - int(date_avant[-2:]) != 1 and int(date[:4]) - int(date_avant[:4]) == 0 :
            for k in range(int(date_avant[-2:])+1, int(date[-2:])):
                if k < 10:
                    dico_a_ajouter[date[:4] + '/0' + str(k)] = valeur_avant
                else :
                    dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
        elif int(date[-2:]) - int(date_avant[-2:]) != 11 and int(date[:4]) - int(date_avant[:4]) == 1:
            for k in range(int(date_avant[-2:])+1, int(date[-2:])+12):
                if k <= 12:
                    if k <10:
                        dico_a_ajouter[date_avant[:4] + '/0' + str(k)] = valeur_avant
                    else:
                        dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
                else :
                    if k%12 < 10:
                        dico_a_ajouter[date[:4] + '/0' + str(k%12)] = valeur_avant
                    else :
                        dico_a_ajouter[date[:4] + '/' + str(k%12)] = valeur_avant

    for el in dico_a_ajouter.items():
        date, valeur = el
        dico_fenetre[date] = valeur
        
    print('Importing prices for material : windows : done')
    
    #garde corps
    print('Importing prices for material : garde corps')

    wb_garde_corps = openpyxl.load_workbook('Data/Extracts prix/Price Development Material Garde Corps.xlsx')
    sheet_garde_corps = wb_garde_corps.active

    max_col_garde_corps = sheet_garde_corps.max_column
    dico_garde_corps = {}
    for i in range(2, max_col_garde_corps+1):
        dico_garde_corps[sheet_garde_corps.cell(row = 7, column = i).value] = sheet_garde_corps.cell(row = 8, column = i).value 

    liste = list(dico_garde_corps.items())
    dico_a_ajouter = {} 
    for el in range(1, len(dico_garde_corps.items())) :
        date, valeur = liste[el]
        date_avant, valeur_avant = liste[el-1]
        if int(date[-2:]) - int(date_avant[-2:]) != 1 and int(date[:4]) - int(date_avant[:4]) == 0 :
            for k in range(int(date_avant[-2:])+1, int(date[-2:])):
                if k < 10:
                    dico_a_ajouter[date[:4] + '/0' + str(k)] = valeur_avant
                else :
                    dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
        elif int(date[-2:]) - int(date_avant[-2:]) != 11 and int(date[:4]) - int(date_avant[:4]) == 1:
            for k in range(int(date_avant[-2:])+1, int(date[-2:])+12):
                if k <= 12:
                    if k <10:
                        dico_a_ajouter[date_avant[:4] + '/0' + str(k)] = valeur_avant
                    else:
                        dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
                else :
                    if k%12 < 10:
                        dico_a_ajouter[date[:4] + '/0' + str(k%12)] = valeur_avant
                    else :
                        dico_a_ajouter[date[:4] + '/' + str(k%12)] = valeur_avant

    for el in dico_a_ajouter.items():
        date, valeur = el
        dico_garde_corps[date] = valeur
        
    print('Importing prices for material : garde corps : done')
    
    #roof
    print('Importing prices for material : roof skeleton')
    wb_toit = openpyxl.load_workbook('Data/Extracts prix/Price Development Material Ossature Toit.xlsx')
    sheet_toit = wb_toit.active

    max_col_toit = sheet_toit.max_column
    dico_toit = {}
    for i in range(2, max_col_toit+1):
        dico_toit[sheet_toit.cell(row = 7, column = i).value] = sheet_toit.cell(row = 8, column = i).value

    liste = list(dico_toit.items())
    dico_a_ajouter = {} 
    for el in range(1, len(dico_toit.items())) :
        date, valeur = liste[el]
        date_avant, valeur_avant = liste[el-1]
        if int(date[-2:]) - int(date_avant[-2:]) != 1 and int(date[:4]) - int(date_avant[:4]) == 0 :
            for k in range(int(date_avant[-2:])+1, int(date[-2:])):
                if k < 10:
                    dico_a_ajouter[date[:4] + '/0' + str(k)] = valeur_avant
                else :
                    dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
        elif int(date[-2:]) - int(date_avant[-2:]) != 11 and int(date[:4]) - int(date_avant[:4]) == 1:
            for k in range(int(date_avant[-2:])+1, int(date[-2:])+12):
                if k <= 12:
                    if k <10:
                        dico_a_ajouter[date_avant[:4] + '/0' + str(k)] = valeur_avant
                    else:
                        dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
                else :
                    if k%12 < 10:
                        dico_a_ajouter[date[:4] + '/0' + str(k%12)] = valeur_avant
                    else :
                        dico_a_ajouter[date[:4] + '/' + str(k%12)] = valeur_avant

    for el in dico_a_ajouter.items():
        date, valeur = el
        dico_toit[date] = valeur
    
    print('Importing prices for material : roof skeleton : done')
    
    #panel
    
    print('Importing prices for material : panel')
    
    wb_panel = openpyxl.load_workbook('Data/Extracts prix/Price Development Panel.xlsx')
    sheet_panel = wb_panel.active

    max_col_panel = sheet_panel.max_column
    dico_panel = {}
    for i in range(2, max_col_panel+1):
        dico_panel[sheet_panel.cell(row = 7, column = i).value] = sheet_panel.cell(row = 8, column = i).value 

    liste = list(dico_panel.items())
    dico_a_ajouter = {} 
    for el in range(1, len(dico_panel.items())) :
        date, valeur = liste[el]
        date_avant, valeur_avant = liste[el-1]
        if int(date[-2:]) - int(date_avant[-2:]) != 1 and int(date[:4]) - int(date_avant[:4]) == 0 :
            for k in range(int(date_avant[-2:])+1, int(date[-2:])):
                if k < 10:
                    dico_a_ajouter[date[:4] + '/0' + str(k)] = valeur_avant
                else :
                    dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
        elif int(date[-2:]) - int(date_avant[-2:]) != 11 and int(date[:4]) - int(date_avant[:4]) == 1:
            for k in range(int(date_avant[-2:])+1, int(date[-2:])+12):
                if k <= 12:
                    if k <10:
                        dico_a_ajouter[date_avant[:4] + '/0' + str(k)] = valeur_avant
                    else:
                        dico_a_ajouter[date_avant[:4] + '/' + str(k)] = valeur_avant
                else :
                    if k%12 < 10:
                        dico_a_ajouter[date[:4] + '/0' + str(k%12)] = valeur_avant
                    else :
                        dico_a_ajouter[date[:4] + '/' + str(k%12)] = valeur_avant

    for el in dico_a_ajouter.items():
        date, valeur = el
        dico_panel[date] = valeur
    
    print('Importing prices for material : panel : done')
    
    #conversion into df and creation of tab_achat
    print('prices file creation')
    
    df_achat_material_fenetre = pd.DataFrame.from_dict(dico_fenetre, orient = 'Index', columns = ['Fenetre [EUR]'])
    df_achat_material_panel = pd.DataFrame.from_dict(dico_panel, orient = 'Index', columns = ['Panel [EUR]'])
    df_achat_material_toit = pd.DataFrame.from_dict(dico_toit, orient = 'Index', columns = ['Ossature toit [EUR]'])
    df_achat_material_garde_corps = pd.DataFrame.from_dict(dico_garde_corps, orient = 'Index', columns = ['Garde corps [EUR]'])

    tab_achat = df_achat_material_fenetre.join(df_achat_material_panel).join(df_achat_material_toit).join(df_achat_material_garde_corps).melt(ignore_index = False).reset_index().rename(columns = {'index':'Date'})
    tab_achat['Date'] = pd.to_datetime(tab_achat.Date)
    tab_achat['Date'] = tab_achat['Date'].dt.strftime('%m/%Y')
    tab_achat = tab_achat.groupby('variable').apply(lambda x: x.ffill().bfill())
    tab_achat.to_csv(r'data_extraite/tableau_achat.csv', index = True, header = True)

    print('prices file creation : done')

if __name__ == '__main__':
    main()
